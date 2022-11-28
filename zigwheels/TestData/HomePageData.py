import openpyxl
class HomePageData:


    @staticmethod
    def getTestData(value):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Admin\\zigwheels\\zigwheelData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == value:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
